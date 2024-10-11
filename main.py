import os.path
import time
from typing import List

import openpyxl

import config
from file_reader import FileReader
from translation import TransactionRecord
from openpyxl.styles import Font, PatternFill


def init_env():
    if not os.path.exists(config.output_dir):
        os.mkdir(config.output_dir)


def get_file_path(file_name):
    return os.path.join(config.output_dir, file_name)


# 获取时间字符串
def get_time():
    return time.strftime("%Y-%m-%d %H：%M", time.localtime())


# 先识别是那个银行的交易流水
def get_bank_translation(file):
    print("start recognition translation type:", file)
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(file)

    try:
        # 获取当前活动工作表（默认为第一张）
        sheet = workbook.active

        # 遍历每一行并输出 A 列的值
        for row in sheet.iter_rows(min_row=1, max_row=config.prefetch_row, values_only=True):
            for cell in row:
                clazz = config.BankTranslation.get(cell)
                if clazz is not None:
                    return clazz

    finally:
        workbook.close()


def is_subset(list1, list2):
    set1 = set(list1)
    set2 = set(list2)
    return set1.issubset(set2)


def get_bank_translation_records(execl_file, tx_class):
    workbook = openpyxl.load_workbook(execl_file)

    try:
        sheet = workbook.active

        # 是不是交易开始行, 在Excel中，行号和列号是从1开始的，而不是从0
        tx_start_row_number = 0

        # 转化为交易对象
        data = []

        for row in sheet.iter_rows(min_row=1):
            row_number, column_count, values, not_empty = read_row(row)
            if is_subset(tx_class.field_list, values):
                tx_start_row_number = row_number

            if tx_start_row_number != 0 and row_number > tx_start_row_number and not_empty:
                # 只需要指定长度的列表项，多的不要
                args = values[:len(tx_class.field_list)]
                tx = tx_class(*args)
                print(tx)
                data.append(tx)

        return data
    finally:
        workbook.close()


def read_row(row):
    row_number = 0
    column_count = 0
    values = []

    for cell in row:
        row_number = cell.row
        column_count += 1
        values.append(cell.value)

    return row_number, column_count, values, values.__getitem__(0) is not None


def write_xlsx(translation_records: List[TransactionRecord]):
    if len(translation_records) == 0:
        return

    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active

    # 将字段写入第一行
    ws.append(TransactionRecord.field_list)

    for tr in translation_records:
        row = [
            tr.date, tr.original_currency_amount, tr.rmb_amount, tr.currency,
            tr.description, tr.payer, tr.payee, tr.revenue_or_expense,
            tr.business_entity, tr.purpose, tr.business_type, tr.posting_status,
            tr.remarks
        ]
        ws.append(row)

    set_style(ws)

    wb.save(get_file_path("汇总表_{}.xlsx".format(get_time())))
    wb.close()


def set_style(ws):
    # 设置字体样式
    font_style_header = Font(name='微软雅黑', size=12, bold=True)  # 第一行字体，12号加粗
    font_style_content = Font(name='微软雅黑', size=12)  # 其余内容字体，12号
    fill_yellow = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')  # 黄色背景

    for cell in ws[1]:  # 第一行的单元格
        cell.font = font_style_header
        cell.fill = fill_yellow

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(TransactionRecord.field_list)):
        for cell in row:
            cell.font = font_style_content  # 设置内容的字体样式


if __name__ == '__main__':
    init_env()

    reader = FileReader("/Users/snlan/py_path/report-tools/data")
    files = reader.read()

    all_tx = []
    for file in files:
        tx_clazz = get_bank_translation(file)
        if tx_clazz is not None:
            txs = get_bank_translation_records(file, tx_clazz)
            for tx in txs:
                all_tx.append(tx.convert())

    write_xlsx(all_tx)
